#############################
##本代码包括四个部分：
##1、解压缩or 选择文件夹
##2、遍历文件；
##3、分析文件；
##4、输出表格
############################
import sys
import tkinter as tk
import zipfile
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import os
import time
from tkinter import simpledialog
from tkinter import filedialog, messagebox, Tk, Toplevel, Label, Button, LEFT, RIGHT  
import chardet  # 用于编码检测
import shutil  # 用于删除文件和文件夹
import json  # 用于读写配置文件
from dateutil.parser import parse ##用于识别更为灵活的日期

config_file = 'user_config.json'

def read_config():
    """读取用户配置"""
    try:
        with open(config_file, 'r') as file:
            return json.load(file)
    except FileNotFoundError:
        return {}

def write_config(config):
    """写入用户配置"""
    with open(config_file, 'w') as file:
        json.dump(config, file)

# 设置环境变量以抑制Tkinter的弃用警告
os.environ['TK_SILENCE_DEPRECATION'] = '1'

def detect_encoding(byte_sequence):
    """检测字节序列的编码"""
    result = chardet.detect(byte_sequence)
    encoding = result['encoding']
    return encoding

def ensure_correct_encoding(file_name):
    """尝试以正确的编码解码文件名"""
    try:
        # 尝试使用utf-8解码
        decoded_name = file_name.encode('utf-8').decode('utf-8')
    except UnicodeDecodeError:
        try:
            # 如果失败，尝试使用系统默认编码
            decoded_name = file_name.encode('cp437').decode()
        except UnicodeDecodeError:
            # 如果仍然失败，保持原样
            decoded_name = file_name
    return decoded_name

def custom_dialog(root):
    """弹出自定义对话框让用户选择处理文件夹或压缩包"""
    dialog = Toplevel(root)
    dialog.title("选择处理类型")
    Label(dialog, text="你想要处理一个文件夹还是一个压缩包？").pack(pady=10)

    # 对话框位置
    dialog_width = 300
    dialog_height = 100
    screen_width = dialog.winfo_screenwidth()
    screen_height = dialog.winfo_screenheight()
    x = (screen_width - dialog_width) / 2
    y = (screen_height - dialog_height) / 2
    dialog.geometry(f'{dialog_width}x{dialog_height}+{int(x)}+{int(y)}')

    user_choice = None  # 局部变量来保存用户选择

    def on_zip():
        nonlocal user_choice
        user_choice = '压缩包'
        dialog.destroy()

    def on_folder():
        nonlocal user_choice
        user_choice = '文件夹'
        dialog.destroy()

    Button(dialog, text="压缩包", command=on_zip).pack(side=LEFT, padx=10, pady=10)
    Button(dialog, text="文件夹", command=on_folder).pack(side=RIGHT, padx=10, pady=10)
    dialog.transient(root)  # Make the dialog on top of the main window
    dialog.grab_set()  # Make the dialog modal
    root.wait_window(dialog)  # Wait for the dialog to close before continue

    return user_choice  # 返回用户的选择

def apply_style_to_excel(ws):##表格文件的样式
    """应用样式到Excel工作表"""
    # 设置首行的样式
    header_font = Font(name='微软雅黑', size=24, bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="6495ED", end_color="6495ED", fill_type="solid")

    for cell in ws["1:1"]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill

    # 应用样式到整个表格中的所有单元格
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    all_alignment = Alignment(wrap_text=True)
    normal_font = Font(size=16)
    
    # 从第2行开始，到倒数第二行
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row-1):
        for cell in row:
            cell.border = thin_border
            cell.alignment = all_alignment

            if cell.column == 2:  # 如果是第二列
            # 只为文档名称添加《》，如果它们还未被添加
               if not (str(cell.value).startswith('《') and str(cell.value).endswith('》')):
                    cell.value = f"《{cell.value}》" #加《》

            cell.font = normal_font
    
    # 设置最后一行的样式
    last_row_font = Font(size=24, bold=True)
    last_row_fill = PatternFill(start_color="D56A6B", end_color="D56A6B", fill_type="solid")
    last_row_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[ws.max_row]:
        cell.font = last_row_font
        cell.alignment = last_row_alignment
        cell.fill = last_row_fill
        cell.border = thin_border

    # 设置列宽
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 40
def process_file_date_about(filename):
    """从文件名中提取日期，如果提取失败，则返回文件的最后修改时间"""
    date_pattern = r'([0-9]{4}[/-][0-9]{1,2}[/-][0-9]{1,2}|[0-9]{1,2}[/-][0-9]{1,2}[/-][0-9]{4})'
    date_match = re.search(date_pattern, filename)

    if date_match:
        try:
            # 尝试解析匹配到的日期字符串
            parsed_date = parse(date_match.group())
            return parsed_date.date().isoformat()
        except ValueError:
            # 如果解析日期出错，则返回None
            return None
    else:
        # 如果没有找到匹配的日期格式，则返回None
        return None

def process_file(root, file_path, file_pattern):
    """解析文件名以获取信息，若无创建日期或工作时长，则通过对话框提醒用户输入"""
    basename = os.path.basename(file_path)
    form_time = work_hours = None
    
    # 定义日期模式
    date_pattern = r'([0-9]{4}[/-][0-9]{1,2}[/-][0-9]{1,2}|[0-9]{1,2}[/-][0-9]{1,2}[/-][0-9]{4})'
    
    # 调用 process_file_date_about 处理日期，获取返回值
    form_time = process_file_date_about(basename)
    
    # 检查文件名中是否存在 date_pattern 定义的日期格式
    date_match = re.search(date_pattern, basename)
    
    if form_time is None and date_match:
        # 如果 process_file_date_about 返回 None 但是在文件名中找到了日期格式
        # 可以在这进行特定处理，比如解析 date_match 获取日期字符串
        pass
    
    # 如果 'form_time' 仍然是 None，使用文件的最后修改时间作为备选日期
    if form_time is None:
        mod_time = time.localtime(os.path.getmtime(file_path))
        form_time = time.strftime("%Y-%m-%d", mod_time)
    
    # 尝试解析文件名获取工作时长
    hours_pattern = r'(\d+\.\d+|\.\d+|\d+)h'
    hours_match = re.search(hours_pattern, basename, re.IGNORECASE)
    
    if hours_match:
        work_hours = float(hours_match.group().rstrip('h'))
    else:
        work_hours = 0  # 如果没有找到合适的工作时长格式，默认为0
    
    # 移除日期和工作时长信息后得到文档名称
    doc_name = basename
    if date_match:
        doc_name = re.sub(date_pattern, '', doc_name)
    if hours_match:
        doc_name = re.sub(hours_pattern, '', doc_name)
    
    doc_name = doc_name.strip(" -_.,")  # 去除文件名中的多余字符
    
    if not doc_name:
        doc_name = "未知文档"

    return (form_time, doc_name, work_hours)

def excel_from_files(files_info):##表格文件
    """从文件信息创建Excel工作簿"""
    wb = Workbook()
    ws = wb.active
    ws.append(["文件最后修改时间", "文档名称", "工作时长"])

    total_hours = 0
    file_count = 0  # 初始化文件数量计数
    for file_info in files_info:
        ws.append(file_info)
        total_hours += file_info[2]  # 累加工作时长
        file_count += 1  # 累加文件数量

    # 添加总计行，在第三列显示总工作时长，在第二列显示文件总数
    ws.append(["总计", file_count, f"{total_hours}h"])
    apply_style_to_excel(ws)
    return wb

def process_directory(directory_path, output_path):##处理文件夹
    """处理指定文件夹内的文件，然后形成Excel表格"""
    # 初始化保存处理后文件信息的列表
    files_info = []
    
    # 遍历目录内的所有文件
    for root_dir, dirs, files in os.walk(directory_path):
        for file_name in files:
            # 忽略隐藏文件和系统文件
            if file_name.startswith('.') or file_name.startswith('__MACOSX') or file_name == 'DS_Store':
                continue

            # 组合成完整的文件路径
            file_path = os.path.join(root_dir, file_name)
            # 调用process_file函数处理文件，并获取文件信息
            file_info = process_file(file_path)
            # 如果得到了有效文件信息，则添加到信息列表中
            if file_info is not None:
                files_info.append(file_info)

    # 如果文件信息列表为空，则提示用户并返回
    if not files_info:
        messagebox.showinfo("提示", "未找到可处理的文件。")
        return

    # 调用函数，将收集到的文件信息转换为Excel工作簿
    wb = excel_from_files(files_info)
    # 保存Excel工作簿到用户指定的输出路径
    wb.save(output_path)
    messagebox.showinfo("完成", f"Excel文件已保存到: {output_path}")

def extract_and_process(root, zip_path, output_path=None):##处理压缩包
    """处理压缩包内的文件，然后形成表格"""
    # 读取并获取用户配置信息
    user_config = read_config()

    # 弹窗提示用户选择解压目标文件夹
    messagebox.showinfo("解压目标文件夹", "请选择一个位置来解压缩压缩包内容。", parent=root)
    extract_path = filedialog.askdirectory(
        title="请选择解压目标文件夹",
        initialdir=user_config.get('last_extract', os.path.expanduser("~")),
        parent=root
    )

    # 如果用户没有选择解压目标文件夹，则退出程序
    if not extract_path:
        messagebox.showinfo("提示", "未选择解压目标文件夹，程序退出。", parent=root)
        return

    # 保存用户选择的解压路径
    user_config['last_extract'] = extract_path
    write_config(user_config)

    # 如果未提供输出路径，则弹窗提示用户选择Excel文件保存位置
    if output_path is None:
        messagebox.showinfo("保存Excel", "请选择一个位置来保存生成的Excel文件。", parent=root)
        output_path = filedialog.asksaveasfilename(
            title="保存Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            parent=root
        )
        # 如果用户取消，则退出程序
        if not output_path:
            messagebox.showinfo("提示", "未选择Excel文件保存位置，程序退出。", parent=root)
            return

    # 解压缩包
    extracted_files = []
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        zip_info_list = zip_ref.infolist()  # 获取压缩包内文件信息列表
        for zip_info in zip_info_list:
            # 文件编码检测与修正
            encoding = detect_encoding(zip_info.filename.encode())
            zip_info.filename = ensure_correct_encoding(zip_info.filename)
            # 解压文件，并记录解压后的文件路径
            zip_ref.extract(zip_info, extract_path)
            extracted_files.append(os.path.join(extract_path, zip_info.filename))

    # 处理解压后的文件，调用process_file函数
    files_info = []
    for file_path in extracted_files:
        if file_path.startswith('.') or file_path.startswith('__MACOSX') or file_path == 'DS_Store':
            continue
        file_info = process_file(root, file_path)
        if file_info is not None:
            files_info.append(file_info)

    # 删除解压出的所有文件
    for file_path in extracted_files:
        os.remove(file_path)

    # 创建Excel工作簿，并保存
    wb = excel_from_files(files_info)
    wb.save(output_path)
    messagebox.showinfo("完成", f"Excel文件已保存到: {output_path}")

def main():##主函数
    # 初始化Tkinter根窗口
    root = Tk()
    # 先更新一下窗口任务，暂时有助于在一些操作系统上正常显示对话框
    ##root.update_idletasks()
    # 然后隐藏根窗口
    ##root.withdraw()

    # 读取并获取用户配置信息
    user_config = read_config()

    # 显示使用指南对话框，且具有记忆功能
    if user_config.get('show_guide', True):
        show_guide = messagebox.askyesno(
            "使用指南",
            "请按格式：'YYYY-MM-DD-文档名称-工作时长h.xxx'命名文件。\n"
            "例: 2024-01-31-质证意见第2版-1h.docx\n"
            "文件将被整理到Excel表格。\n"
            "未命名文件会提示输入名称。\n"
            "文件未标记工作时长会提示输入时长。\n"
            "未标记日期文件将采用修改日期。\n"
            "是否再次显示本指南？",
            parent=root
        )
        # 根据用户选择更新配置信息
        user_config['show_guide'] = not show_guide
        write_config(user_config)

    # 自定义对话框让用户选择处理文件夹或压缩包
    user_choice = custom_dialog(root)
    print(user_choice) 

    # 用户选择处理文件夹
    if user_choice == '文件夹':
        directory_path = filedialog.askdirectory(
            title="选择文件夹",
            initialdir=user_config.get('last_directory', os.path.expanduser("~")),
            parent=root
        )
        if directory_path:
            user_config['last_directory'] = directory_path
            write_config(user_config)
            output_path = filedialog.asksaveasfilename(
                title="保存Excel文件",
                initialdir=directory_path,
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                parent=root
            )
            if output_path:
                process_directory(directory_path, output_path)
            else:
                messagebox.showinfo("提示", "未选择输出文件，操作已取消。", parent=root)
        else:
            messagebox.showinfo("提示", "未选择文件夹，操作已取消。", parent=root)

    # 用户选择处理压缩包
    elif user_choice == '压缩包':
        zip_path = filedialog.askopenfilename(
            title="选择压缩包",
            initialdir=user_config.get('last_zip', os.path.expanduser("~")),
            filetypes=[("Zip files", "*.zip")],
            parent=root
        )
        if zip_path:
            user_config['last_zip'] = zip_path
            write_config(user_config)
            extract_path = filedialog.askdirectory(
                title="解压缩包至",
                initialdir=user_config.get('last_extract', os.path.expanduser("~")),
                parent=root
            )
            if extract_path:
                user_config['last_extract'] = extract_path
                write_config(user_config)
                output_path = filedialog.asksaveasfilename(
                    title="保存Excel文件",
                    initialdir=os.path.dirname(zip_path),
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx")],
                    parent=root
                )
                if output_path:
                    extract_and_process(zip_path, extract_path, output_path)
                else:
                    messagebox.showinfo("提示", "未选择输出文件，操作已取消。", parent=root)
            else:
                messagebox.showinfo("提示", "未选择临时解压目录，操作已取消。", parent=root)
        else:
            messagebox.showinfo("提示", "未选择压缩包，操作已取消。", parent=root)

    # 销毁Tkinter根窗口
    root.destroy()

if __name__ == "__main__":
    main()
    

    

