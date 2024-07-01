import random
import itertools
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# 定义队伍和种子队
teams = [
    "水果沙拉队", "疯狂星期四队", "乐观王国队", "DK&YQS海军四大将队", "刚铎队", "落汗队", "X1队", "DK随机组合队", 
    "DK剑盾兄弟团", "KTP报名表为什么设置的这么繁琐队", "DK神猪无敌小偷队", "DK小伙子和老朋友队", "神威无敌杰森大将军队", 
    "圣殿哈哈队", "DK红薯骑士队", "DK精锐队", "DK凶巴巴的老爷队", "DK瘦巴巴的老爷队", "MINI队", "卢管大队"
]
seed_teams = ["X1队", "DK小伙子和老朋友队", "DK&YQS海军四大将队", "MINI队"]
special_restrictions = [("圣殿哈哈队", "X1队")]

# 定义分组
random.seed(42)
groups = {"Group-A": [], "Group-B": [], "Group-C": [], "Group-D": []}

# 分配种子队
for i, seed_team in enumerate(seed_teams):
    groups[f"Group-{chr(65 + i)}"].append(seed_team)

remaining_teams = [team for team in teams if team not in seed_teams]

# 确保圣殿哈哈队和X1队不在同一组
special_team_1, special_team_2 = special_restrictions[0]
for group_name in groups:
    if special_team_1 not in groups[group_name] and special_team_2 not in groups[group_name]:
        groups[group_name].append(special_team_1)
        remaining_teams.remove(special_team_1)
        break

# 随机分配剩余队伍
for team in remaining_teams:
    group_name = random.choice(list(groups.keys()))
    while len(groups[group_name]) >= 5:
        group_name = random.choice(list(groups.keys()))
    groups[group_name].append(team)

# 定义比赛时间模板
match_schedule_template = [
    "周五晚8点场", "周五晚10点场", "周六下午", 
    "周六晚7点场", "周六晚8点场", "周天下午", 
    "周天晚7点场", "周天晚8点场"
]

# 生成每组的对阵表
def generate_round_robin_matches(teams):
    matches = []
    for team1, team2 in itertools.combinations(teams, 2):
        matches.append((team1, team2))
    return matches

# 创建工作簿和样式
wb = Workbook()
header_font = Font(bold=True, size=14)
header_fill = PatternFill("solid", fgColor="DDEBF7")
center_alignment = Alignment(horizontal="center", vertical="center")

# 创建每个组的工作表
for group_name, team_list in groups.items():
    ws = wb.create_sheet(title=group_name)
    
    # 写入分组及积分表
    ws.append(["分组及积分表"])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    
    points_headers = ["队伍名称", "胜", "平", "负", "小局得分", "小局失分", "净胜分", "积分"]
    ws.append(points_headers)
    for cell in ws[2]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    for team in team_list:
        ws.append([team, 0, 0, 0, 0, 0, 0, 0])
    
    # 设置行高
    for row in ws.iter_rows(min_row=1, max_row=len(team_list) + 2):
        for cell in row:
            cell.alignment = center_alignment
        ws.row_dimensions[row[0].row].height = 35
    
    # 生成并写入对阵表
    start_row = len(team_list) + 4
    ws.append(["对阵表"])
    for cell in ws[start_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
    
    match_headers = ["对阵情况", "队伍1得分", "队伍2得分", "队伍1", "队伍2", "对阵时间"]
    ws.append(match_headers)
    for cell in ws[start_row + 1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    matches = generate_round_robin_matches(team_list)
    match_schedule = []
    for match_num, (team1, team2) in enumerate(matches):
        match_time = match_schedule_template[match_num % len(match_schedule_template)]
        match_schedule.append({
            "对阵情况": f"{team1} vs {team2}",
            "队伍1得分": "",
            "队伍2得分": "",
            "队伍1": team1,
            "队伍2": team2,
            "对阵时间": match_time
        })
    match_schedule = sorted(match_schedule, key=lambda x: match_schedule_template.index(x["对阵时间"]))
    for row_num, match in enumerate(match_schedule):
        ws.append([match["对阵情况"], match["队伍1得分"], match["队伍2得分"], match["队伍1"], match["队伍2"], match["对阵时间"]])
    
    # 设置列宽
    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 20
    
    # 添加公式
    for row_num in range(3, 3 + len(team_list)):
        team_name = ws.cell(row=row_num, column=1).value
        win_formula = f'=SUMPRODUCT((D${start_row + 2}:D${start_row + 1 + len(matches)}="{team_name}")*(B${start_row + 2}:B${start_row + 1 + len(matches)}<>"")*(C${start_row + 2}:C${start_row + 1 + len(matches)}<>"")*(B${start_row + 2}:B${start_row + 1 + len(matches)}>C${start_row + 2}:C${start_row + 1 + len(matches)})) + SUMPRODUCT((E${start_row + 2}:E${start_row + 1 + len(matches)}="{team_name}")*(B${start_row + 2}:B${start_row + 1 + len(matches)}<>"")*(C${start_row + 2}:C${start_row + 1 + len(matches)}<>"")*(C${start_row + 2}:C${start_row + 1 + len(matches)}>B${start_row + 2}:B${start_row + 1 + len(matches)}))'
        draw_formula = f'=SUMPRODUCT((D${start_row + 2}:D${start_row + 1 + len(matches)}="{team_name}")*(B${start_row + 2}:B${start_row + 1 + len(matches)}=C${start_row + 2}:C${start_row + 1 + len(matches)})*(B${start_row + 2}:B${start_row + 1 + len(matches)}<>"")) + SUMPRODUCT((E${start_row + 2}:E${start_row + 1 + len(matches)}="{team_name}")*(B${start_row + 2}:B${start_row + 1 + len(matches)}=C${start_row + 2}:C${start_row + 1 + len(matches)})*(B${start_row + 2}:B${start_row + 1 + len(matches)}<>""))'
        lose_formula = f'=SUMPRODUCT((D${start_row + 2}:D${start_row + 1 + len(matches)}="{team_name}")*(B${start_row + 2}:B${start_row + 1 + len(matches)}<>"")*(C${start_row + 2}:C${start_row + 1 + len(matches)}<>"")*(B${start_row + 2}:B${start_row + 1 + len(matches)}<C${start_row + 2}:C${start_row + 1 + len(matches)})) + SUMPRODUCT((E${start_row + 2}:E${start_row + 1 + len(matches)}="{team_name}")*(B${start_row + 2}:B${start_row + 1 + len(matches)}<>"")*(C${start_row + 2}:C${start_row + 1 + len(matches)}<>"")*(C${start_row + 2}:C${start_row + 1 + len(matches)}<B${start_row + 2}:B${start_row + 1 + len(matches)}))'
        points_for_formula = f'=SUMIF(D${start_row + 2}:D${start_row + 1 + len(matches)}, "{team_name}", B${start_row + 2}:B${start_row + 1 + len(matches)}) + SUMIF(E${start_row + 2}:E${start_row + 1 + len(matches)}, "{team_name}", C${start_row + 2}:C${start_row + 1 + len(matches)})'
        points_against_formula = f'=SUMIF(D${start_row + 2}:D${start_row + 1 + len(matches)}, "{team_name}", C${start_row + 2}:C${start_row + 1 + len(matches)}) + SUMIF(E${start_row + 2}:E${start_row + 1 + len(matches)}, "{team_name}", B${start_row + 2}:B${start_row + 1 + len(matches)})'
        net_points_formula = f'=E{row_num} - F{row_num}'
        total_points_formula = f'=B{row_num}*3 + C{row_num}*1'
        ws.cell(row=row_num, column=2).value = win_formula
        ws.cell(row=row_num, column=3).value = draw_formula
        ws.cell(row=row_num, column=4).value = lose_formula
        ws.cell(row=row_num, column=5).value = points_for_formula
        ws.cell(row=row_num, column=6).value = points_against_formula
        ws.cell(row=row_num, column=7).value = net_points_formula
        ws.cell(row=row_num, column=8).value = total_points_formula

# 保存工作簿
wb.save("tournament_schedule_v14.xlsx")

# 生成VBA脚本
vba_script = """
Private Sub Worksheet_Change(ByVal Target As Range)
    Dim KeyCells As Range
    Set KeyCells = Range("B:H")
    
    If Not Application.Intersect(KeyCells, Range(Target.Address)) Is Nothing Then
        Call SortRanking
    End If
End Sub

Sub SortRanking()
    Dim ws As Worksheet
    Dim lastRow As Long
    
    For Each ws In ThisWorkbook.Worksheets
        If ws.Name Like "Group-*" Then
            lastRow = ws.Cells(ws.Rows.Count, "A").End(xlUp).Row
            With ws.Sort
                .SortFields.Clear
                .SortFields.Add Key:=ws.Range("H3:H" & lastRow), Order:=xlDescending
                .SortFields.Add Key:=ws.Range("E3:E" & lastRow), Order:=xlDescending
                .SortFields.Add Key:=ws.Range("G3:G" & lastRow), Order:=xlDescending
                .SortFields.Add Key:=ws.Range("F3:F" & lastRow), Order:=xlAscending
                .SetRange ws.Range("A2:H" & lastRow)
                .Header = xlYes
                .Apply
            End With
        End If
    Next ws
End Sub
"""

# 将VBA脚本添加到工作簿
vba_module_name = "SortRanking"
with open("SortRanking.bas", "w") as f:
    f.write(vba_script)

print("Excel文件和VBA脚本生成完毕。")
