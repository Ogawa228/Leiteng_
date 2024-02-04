import sys
import zipfile
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment,Font
from openpyxl.worksheet.datavalidation import DataValidation
import os
import time
from tkinter import simpledialog
from tkinter import filedialog, messagebox, Tk
from datetime import datetime
from tkinter import Toplevel, Label, Button, LEFT, RIGHT, Tk #自定义对话框
import chardet  # 用于编码检测
import shutil  # 用于删除文件和文件夹
import json  # 用于读写配置文件

config_file = 'user_config.json'

def read_config():
    """读取用户配置"""
    try:
        with open(config_file, 'r') as file:
            return json.load(file)
    except FileNotFoundError:
        return {}

def write_config(config):
    """写入用户配置"""
    with open(config_file, 'w') as file:
        json.dump(config, file)



# 设置环境变量以抑制Tkinter的弃用警告
os.environ['TK_SILENCE_DEPRECATION'] = '1'
# 弹出对话框指引用户使用程序
messagebox.showinfo
(
     "雷腾小伙伴减负指南",

            "你可以将压缩包或者文件夹内文件命名为以下格式：\n"
            "\n"
            "YYYY-MM-DD-文档名称-工作时长h.xxx\n"
            "例如:2024-01-31-质证意见第2版-1h.docx\n"
            "\n"
            "所有文件会被整理到表格中\n"
            "没写工作时长的会提示你填\n"            
            "没写时间的会直接抓取最后更改时间\n"
            "\n"
)

root = Tk()
root.withdraw()  # 不显示Tk根窗口


def custom_dialog(root):
    dialog = Toplevel(root)
    dialog.title("选择处理类型")
    Label(dialog, text="你想要处理一个文件夹还是一个压缩包？").pack(pady=10)
    
    def on_zip():
        global user_choice
        user_choice = '压缩包'
        dialog.destroy()
    
    def on_folder():
        global user_choice
        user_choice = '文件夹'
        dialog.destroy()
    
    Button(dialog, text="压缩包", command=on_zip).pack(side=LEFT, padx=10, pady=10)
    Button(dialog, text="文件夹", command=on_folder).pack(side=RIGHT, padx=10, pady=10)
    dialog.transient(root)  # Make the dialog on top of the main window
    dialog.grab_set()  # Make the dialog modal
    root.wait_window(dialog)  # Wait for the dialog to close before continue

root = Tk()
root.withdraw()

def detect_encoding(byte_sequence):
    """检测字节序列的编码"""
    result = chardet.detect(byte_sequence)
    encoding = result['encoding']
    return encoding

def ensure_correct_encoding(file_name):
    """尝试以正确的编码解码文件名"""
    try:
        # 尝试使用utf-8解码
        decoded_name = file_name.encode('utf-8').decode('utf-8')
    except UnicodeDecodeError:
        try:
            # 如果失败，尝试使用系统默认编码
            decoded_name = file_name.encode('cp437').decode()
        except UnicodeDecodeError:
            # 如果仍然失败，保持原样
            decoded_name = file_name
    return decoded_name

    """处理指定文件夹内的文件，然后形成表格"""
    # 初始化Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.append(["文件最后修改时间", "文档名称", "工作时长"])
    total_hours = 0

    # 遍历解压后的文件进行处理
    for root, dirs, files in os.walk(extract_path):
        for file in files:
            print("Processing file:", file)
            # 跳过系统文件和隐藏文件
            if file.startswith('.') or file.startswith('__MACOSX') or file == 'DS_Store':
                continue

             # 尝试从文件名中提取数据
            match = re.match(r"(\d{4}-\d{1,2}-\d{1,2})-([\u4e00-\u9fa5]+)-(\d+(\.\d)?)h\.(docx|doc|xlsx|xls)", file)
            if match:
                form_time, doc_name, work_hours, _, _ = match.groups()
                work_hours = float(work_hours)
                doc_name = f"《{doc_name}》"
            else:
                # 尝试从文件名中提取日期和汉字内容
                date_match = re.match(r"(\d{4}-\d{1,2}-\d{1,2})", file)
                name_match = re.search(r"([\u4e00-\u9fa5]+)", file)
                if date_match:
                    form_time = date_match.group(1)
                else:
                    mod_time = time.localtime(os.path.getmtime(os.path.join(root, file)))
                    form_time = time.strftime("%Y-%m-%d", mod_time)
                doc_name = f"《{name_match.group(1)}》" if name_match else "《未知文档》"

                # 尝试从文件名中提取工作时长，如果没有则询问用户
                work_hours_match = re.search(r"(\d+(\.\d)?)h", file)
                if work_hours_match:
                    work_hours = float(work_hours_match.group(1))
                else:
                    work_hours = simpledialog.askfloat("输入工作时长", f"请输入文件'{file}'的工作时长（小时）:", minvalue=0.0)
                    if work_hours is None:  # 用户取消输入
                        continue

            total_hours += work_hours
            ws.append([form_time, doc_name, f"{work_hours}h"])

# 添加总计行
    ws.append(["总计", "", f"{total_hours}h"])

    # 获取所有行的迭代器
    rows = list(ws.iter_rows())

    # 获取最后一行，即总计行
    total_row = rows[-1]

    # 设置第二列的计算总文件的个数
    total_row[1].font = Font(bold=True)
    total_row[1].fill = PatternFill(start_color="CD5C5C", end_color="CD5C5C", fill_type="solid")
    total_row[1].alignment = Alignment(horizontal="center", vertical="center")

    # 设置第三列的计算总工作时长
    total_row[2].font = Font(bold=True)
    total_row[2].fill = PatternFill(start_color="CD5C5C", end_color="CD5C5C", fill_type="solid")
    total_row[2].alignment = Alignment(horizontal="center", vertical="center")

    # 应用样式到整个工作表
    apply_style_to_excel(ws)
    adjust_column_width_and_row_height(ws)

    # 最后保存工作簿
    wb.save(output_path)



    if output_path is None:
        output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not output_path:  # 如果用户取消选择
            print("未选择输出文件，程序退出。")
            sys.exit(1)

    wb.save(output_path)
    print(f"Excel文件已保存到: {output_path}")

def apply_style_to_excel(ws):
    """应用样式到Excel工作表"""
    # 设置首行的样式
    header_font = Font(name='微软雅黑', size=24, bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="6495ED", end_color="6495ED", fill_type="solid")

    for cell in ws["1:1"]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill

    # 设置文档名称列（第二列）的样式：加粗、居中，字体大小16
    doc_name_alignment = Alignment(horizontal="center", vertical="center")
    for cell in ws['B']:
        if cell.row == 1:  # 首行
            cell.font = Font(name='微软雅黑', size=24, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:  # 除首行外的其它行
            cell.font = Font(name='微软雅黑', size=16, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 设置边框和自动换行
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=2): #从第2行开始
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)

def adjust_column_width_and_row_height(ws):
    """设置固定列宽并根据内容调整行高"""
    # 设置固定列宽，大约每厘米7个字符宽度
    ws.column_dimensions['A'].width = 5 * 7
    ws.column_dimensions['B'].width = 13 * 7
    ws.column_dimensions['C'].width = 3.5 * 7

    # 设置最小行高为3cm对应的磅值
    min_height_pt = 3 * 28.35  # 3cm in points

    for row in ws.iter_rows():
        max_length = max(len(str(cell.value)) for cell in row)
        # 根据最长的单元格内容设置行高，确保最小行高为3cm
        ws.row_dimensions[row[0].row].height = max(min_height_pt, max_length * 1.2)


def add_data_validation(ws):
    """在Excel工作表中添加数据有效性筛查"""
    # 文件最后修改时间的数据有效性检查（日期格式）
    date_validation = DataValidation(type="date", formula1='"1900-01-01"', formula2='"9999-12-31"', operator="between", allow_blank=True)
    ws.add_data_validation(date_validation)
    date_validation.add(ws['A2:A1048576'])  # 假设日期在第一列

    # 文档名称的数据有效性检查（仅汉字）
    # 注意：Excel本身不支持直接对汉字进行数据有效性检查，此处仅为示例
    # text_validation = DataValidation(type="custom", formula1='...')  # 需要自定义公式
    # ws.add_data_validation(text_validation)
    # text_validation.add(ws['B2:B1048576'])  # 假设文档名称在第二列

    #工作时长的数据有效性检查（数字）
    time_validation = DataValidation(type="decimal", formula1="0", formula2="99999.9", operator="between", allow_blank=True)
    ws.add_data_validation(time_validation)
    time_validation.add(ws['C2:C1048576'])  #假设工作时长在第三列

def extract_and_process(zip_path, output_path=None):
    messagebox.showinfo("选择解压目标文件夹", "请选择要将临时内容解压到哪个文件夹。")
    extract_path = filedialog.askdirectory(title="请选择解压目标文件夹")
    if not extract_path:
        print("未选择解压目标文件夹，程序退出。")
        sys.exit(1)

    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        for member in zip_ref.infolist():
            original_file_name = member.filename
            file_name_utf8 = ensure_correct_encoding(original_file_name)
            target_path = os.path.join(extract_path, file_name_utf8)
            target_dir = os.path.dirname(target_path)
            if not os.path.exists(target_dir):
                os.makedirs(target_dir)
            if not member.is_dir():
                with zip_ref.open(member) as source_file, open(target_path, 'wb') as target_file:
                    target_file.write(source_file.read())

    #初始化Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.append(["文件最后修改时间", "文档名称", "工作时长"])
    total_hours = 0

    # 遍历解压后的文件进行处理
    for root, dirs, files in os.walk(extract_path):
        for file in files:
            print("Processing file:", file)
            # 跳过系统文件和隐藏文件
            if file.startswith('.') or file.startswith('__MACOSX') or file == 'DS_Store':
                continue

             # 尝试从文件名中提取数据
            match = re.match(r"(\d{4}-\d{1,2}-\d{1,2})-([\u4e00-\u9fa5]+)-(\d+(\.\d)?)h\.(docx|doc|xlsx|xls)", file)
            if match:
                form_time, doc_name, work_hours, _, _ = match.groups()
                work_hours = float(work_hours)
                doc_name = f"《{doc_name}》"
            else:
                # 尝试从文件名中提取日期和汉字内容
                date_match = re.match(r"(\d{4}-\d{1,2}-\d{1,2})", file)
                name_match = re.search(r"([\u4e00-\u9fa5]+)", file)
                if date_match:
                    form_time = date_match.group(1)
                else:
                    mod_time = time.localtime(os.path.getmtime(os.path.join(root, file)))
                    form_time = time.strftime("%Y-%m-%d", mod_time)
                doc_name = f"《{name_match.group(1)}》" if name_match else "《未知文档》"

                # 尝试从文件名中提取工作时长，如果没有则询问用户
                work_hours_match = re.search(r"(\d+(\.\d)?)h", file)
                if work_hours_match:
                    work_hours = float(work_hours_match.group(1))
                else:
                    work_hours = simpledialog.askfloat("输入工作时长", f"请输入文件'{file}'的工作时长（小时）:", minvalue=0.0)
                    if work_hours is None:  # 用户取消输入
                        continue

            total_hours += work_hours
            ws.append([form_time, doc_name, f"{work_hours}h"])

# 添加总计行
    ws.append(["总计", "", f"{total_hours}h"])

    # 获取所有行的迭代器
    rows = list(ws.iter_rows())

    # 获取最后一行，即总计行
    total_row = rows[-1]

    # 设置第二列的计算总文件的个数
    total_row[1].font = Font(bold=True)
    total_row[1].fill = PatternFill(start_color="CD5C5C", end_color="CD5C5C", fill_type="solid")
    total_row[1].alignment = Alignment(horizontal="center", vertical="center")

    # 设置第三列的计算总工作时长
    total_row[2].font = Font(bold=True)
    total_row[2].fill = PatternFill(start_color="CD5C5C", end_color="CD5C5C", fill_type="solid")
    total_row[2].alignment = Alignment(horizontal="center", vertical="center")

    # 应用样式到整个工作表
    apply_style_to_excel(ws)
    adjust_column_width_and_row_height(ws)

    # 最后保存工作簿
    wb.save(output_path)



    if output_path is None:
        output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not output_path:  # 如果用户取消选择
            print("未选择输出文件，程序退出。")
            sys.exit(1)

    wb.save(output_path)
    print(f"Excel文件已保存到: {output_path}")
    # 删除解压出的文件
    for root, dirs, files in os.walk(extract_path):
        for file in files:
            os.remove(os.path.join(root, file))

def process_directory(directory_path, output_path=None):
    """处理指定文件夹内的文件，然后形成表格"""
    # 初始化Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.append(["文件最后修改时间", "文档名称", "工作时长"])
    total_hours = 0

    # 遍历解压后的文件进行处理
    for root, dirs, files in os.walk(extract_path):
        for file in files:
            print("Processing file:", file)
            # 跳过系统文件和隐藏文件
            if file.startswith('.') or file.startswith('__MACOSX') or file == 'DS_Store':
                continue

             # 尝试从文件名中提取数据
            match = re.match(r"(\d{4}-\d{1,2}-\d{1,2})-([\u4e00-\u9fa5]+)-(\d+(\.\d)?)h\.(docx|doc|xlsx|xls)", file)
            if match:
                form_time, doc_name, work_hours, _, _ = match.groups()
                work_hours = float(work_hours)
                doc_name = f"《{doc_name}》"
            else:
                # 尝试从文件名中提取日期和汉字内容
                date_match = re.match(r"(\d{4}-\d{1,2}-\d{1,2})", file)
                name_match = re.search(r"([\u4e00-\u9fa5]+)", file)
                if date_match:
                    form_time = date_match.group(1)
                else:
                    mod_time = time.localtime(os.path.getmtime(os.path.join(root, file)))
                    form_time = time.strftime("%Y-%m-%d", mod_time)
                doc_name = f"《{name_match.group(1)}》" if name_match else "《未知文档》"

                # 尝试从文件名中提取工作时长，如果没有则询问用户
                work_hours_match = re.search(r"(\d+(\.\d)?)h", file)
                if work_hours_match:
                    work_hours = float(work_hours_match.group(1))
                else:
                    work_hours = simpledialog.askfloat("输入工作时长", f"请输入文件'{file}'的工作时长（小时）:", minvalue=0.0)
                    if work_hours is None:  # 用户取消输入
                        continue

            total_hours += work_hours
            ws.append([form_time, doc_name, f"{work_hours}h"])

# 添加总计行
    ws.append(["总计", "", f"{total_hours}h"])

    # 获取所有行的迭代器
    rows = list(ws.iter_rows())

    # 获取最后一行，即总计行
    total_row = rows[-1]

    # 设置第二列的计算总文件的个数
    total_row[1].font = Font(bold=True)
    total_row[1].fill = PatternFill(start_color="CD5C5C", end_color="CD5C5C", fill_type="solid")
    total_row[1].alignment = Alignment(horizontal="center", vertical="center")

    # 设置第三列的计算总工作时长
    total_row[2].font = Font(bold=True)
    total_row[2].fill = PatternFill(start_color="CD5C5C", end_color="CD5C5C", fill_type="solid")
    total_row[2].alignment = Alignment(horizontal="center", vertical="center")

    # 应用样式到整个工作表
    apply_style_to_excel(ws)
    adjust_column_width_and_row_height(ws)

    # 最后保存工作簿
    wb.save(output_path)



    if output_path is None:
        output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not output_path:  # 如果用户取消选择
            print("未选择输出文件，程序退出。")
            sys.exit(1)

    wb.save(output_path)
    print(f"Excel文件已保存到: {output_path}")

if __name__ == "__main__":
    root = Tk()
    root.withdraw()  # 不显示Tk根窗口

    user_config = read_config()

    # 显示使用指南对话框
    if user_config.get('show_guide', True):
        show_guide = messagebox.askyesno
        (
            "雷腾小伙伴减负指南",

            "你可以将压缩包或者文件夹内文件命名为以下格式：\n"
            "\n"
            "YYYY-MM-DD-文档名称-工作时长h.xxx\n"
            "例如:2024-01-31-质证意见第2版-1h.docx\n"
            "\n"
            "所有文件会被整理到表格中\n"
            "没写工作时长的会提示你填\n"            
            "没写时间的会直接抓取最后更改时间\n"
            "\n"
            "下次运行程序时是否还要显示这条消息？"
        )
        user_config['show_guide'] = show_guide
        write_config(user_config)
root = Tk()
root.withdraw()

    # 下面调用自定义的对话框
user_choice = None
custom_dialog(root)

# 根据用户选择处理
if user_choice == '文件夹':
    messagebox.showinfo("选择文件夹", "请选择需要处理的文件夹。")
    directory_path = filedialog.askdirectory()
    if not directory_path:
        sys.exit("未选择文件夹，程序退出。")
    process_directory(directory_path, output_path)  # 假设这是你已经定义好的函数
elif user_choice == '压缩包':
    messagebox.showinfo("选择压缩包", "请选择需要处理的压缩包文件。")
    zip_path = filedialog.askopenfilename(
        title="请选择压缩包",
        initialdir=os.path.expanduser("~"),
        filetypes=[("Zip files", "*.zip")]
    )
    if not zip_path:
        sys.exit("未选择压缩包，程序退出。")
    extract_and_process(zip_path, output_path)  # 假设这是你已经定义好的函数
else:
    sys.exit("用户取消操作，程序退出。")
    root = Tk()
    root.withdraw()
    
    # 使用用户上次选择的路径作为默认路径
    default_path = user_config.get('default_path', os.path.join(os.path.expanduser("~"), "Desktop"))

    # 选择要上传的压缩包
    messagebox.showinfo("选择压缩包", "请选择需要处理的压缩包文件。")
    zip_path = filedialog.askopenfilename(
        title="请选择压缩包",
        initialdir=os.path.expanduser("~"),
        filetypes=[("Zip files", "*.zip")]
    )
    if not zip_path:
        sys.exit("未选择压缩包，程序退出。")
 

    # 选择保存Excel文件的位置
    messagebox.showinfo("保存Excel文件", "请选择输出Excel文件的保存位置。")
    output_path = filedialog.asksaveasfilename(
        title="保存Excel文件",
        initialdir=default_path,
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not output_path:
        sys.exit("未选择输出文件，程序退出。")

    extract_and_process(zip_path, output_path)
