import sys
import zipfile
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment,Font
from openpyxl.worksheet.datavalidation import DataValidation
import os
from tkinter import filedialog, messagebox, Tk
from datetime import datetime
import chardet  # 用于编码检测
import shutil  # 用于删除文件和文件夹
import json  # 用于读写配置文件

config_file = 'user_config.json'

def read_config():
    """读取用户配置"""
    try:
        with open(config_file, 'r') as file:
            return json.load(file)
    except FileNotFoundError:
        return {}

def write_config(config):
    """写入用户配置"""
    with open(config_file, 'w') as file:
        json.dump(config, file)



# 设置环境变量以抑制Tkinter的弃用警告
os.environ['TK_SILENCE_DEPRECATION'] = '1'
# 弹出对话框指引用户使用程序
messagebox.showinfo(
    "雷腾小伙伴减负指南",
    "请将压缩包内文件命名为以下格式：\n"
    "YYYY-MM-DD-文档名称-工作时长h.xxx\n"
    "例如:2024-01-31-质证意见第2版-1h.docx\n"
    "所有文件将被解压并整理到表格中，不符合命名规则的文件将被忽略。"
)

root = Tk()
root.withdraw()  # 不显示Tk根窗口

def detect_encoding(byte_sequence):
    """检测字节序列的编码"""
    result = chardet.detect(byte_sequence)
    encoding = result['encoding']
    return encoding

def ensure_correct_encoding(file_name):
    """尝试以正确的编码解码文件名"""
    try:
        byte_sequence = file_name.encode('cp437')
        detected_encoding = detect_encoding(byte_sequence)
        if detected_encoding in ['GB2312', 'GBK', 'GB18030']:
            return byte_sequence.decode('gbk')
        return byte_sequence.decode(detected_encoding)
    except Exception as e:
        print(f"Error decoding filename: {e}")
        return file_name

def apply_style_to_excel(ws):
    """应用样式到Excel工作表"""
    # 设置首行的样式
    header_font = Font(name='楷体_GB2312',size=24, bold=True)  # 设置字体为楷体，加粗
    header_alignment = Alignment(horizontal="center", vertical="center")  # 水平和垂直居中
    header_fill = PatternFill(start_color="6495ED", end_color="6495ED", fill_type="solid")  # 背景填充

    for cell in ws["1:1"]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill

    # 设置第二列除首行外的字体大小为16
    col_font = Font(name='楷体_GB2312', size=16)  # 设置字体为楷体，字号16
    for cell in ws['B']:
        if cell.row != 1:  # 跳过首行
            cell.font = col_font

    # 设置边框和自动换行
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=2):  # 从第二行开始应用边框和自动换行，保持第一行的样式
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)

def adjust_column_width_and_row_height(ws):
    """设置固定列宽并根据内容调整行高"""
    # 设置固定列宽，大约每厘米7个字符宽度
    ws.column_dimensions['A'].width = 3.5 * 7
    ws.column_dimensions['B'].width = 13 * 7
    ws.column_dimensions['C'].width = 3.5 * 7

    # 设置最小行高为3cm对应的磅值
    min_height_pt = 3 * 28.35  # 3cm in points

    for row in ws.iter_rows():
        max_length = max(len(str(cell.value)) for cell in row)
        # 根据最长的单元格内容设置行高，确保最小行高为3cm
        ws.row_dimensions[row[0].row].height = max(min_height_pt, max_length * 1.2)


def add_data_validation(ws):
    """在Excel工作表中添加数据有效性筛查"""
    # 形成时间的数据有效性检查（日期格式）
    date_validation = DataValidation(type="date", formula1='"1900-01-01"', formula2='"9999-12-31"', operator="between", allow_blank=True)
    ws.add_data_validation(date_validation)
    date_validation.add(ws['A2:A1048576'])  # 假设日期在第一列

    # 文档名称的数据有效性检查（仅汉字）
    # 注意：Excel本身不支持直接对汉字进行数据有效性检查，此处仅为示例
    # text_validation = DataValidation(type="custom", formula1='...')  # 需要自定义公式
    # ws.add_data_validation(text_validation)
    # text_validation.add(ws['B2:B1048576'])  # 假设文档名称在第二列

    #工作时长的数据有效性检查（数字）
    time_validation = DataValidation(type="decimal", formula1="0", formula2="99999.9", operator="between", allow_blank=True)
    ws.add_data_validation(time_validation)
    time_validation.add(ws['C2:C1048576'])  #假设工作时长在第三列

def extract_and_process(zip_path, output_path=None):
    messagebox.showinfo("选择解压目标文件夹", "请选择要将临时内容解压到哪个文件夹。")
    extract_path = filedialog.askdirectory(title="请选择解压目标文件夹")
    if not extract_path:
        print("未选择解压目标文件夹，程序退出。")
        sys.exit(1)

    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        for member in zip_ref.infolist():
            original_file_name = member.filename
            file_name_utf8 = ensure_correct_encoding(original_file_name)
            target_path = os.path.join(extract_path, file_name_utf8)
            target_dir = os.path.dirname(target_path)
            if not os.path.exists(target_dir):
                os.makedirs(target_dir)
            if not member.is_dir():
                with zip_ref.open(member) as source_file, open(target_path, 'wb') as target_file:
                    target_file.write(source_file.read())

    #初始化Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.append(["形成时间", "文档名称", "工作时长"])
    total_hours = 0

    # 遍历解压后的文件进行处理
    for root, dirs, files in os.walk(extract_path):
        for file in files:
            # 跳过系统文件和隐藏文件
            if file.startswith('.') or file.startswith('__MACOSX') or file == 'DS_Store':
                continue

            # 尝试从文件名中提取数据
            match = re.match(r"(\d{4}-\d{1,2}-\d{1,2})-([\u4e00-\u9fa5]+)-(\d+(\.\d)?)h\..+", file)
            if match:
                form_time, doc_name, work_hours, _ = match.groups()
                work_hours_float = float(work_hours)  # 转换为浮点数并保留1位小数
                total_hours += work_hours_float
                ws.append([form_time, doc_name, f"{work_hours_float:.1f}h"])
            else:
                # 如果文件名不符合预期格式，则跳过
                continue

#添加总计行
    ws.append(["总计", "", f"{total_hours}h"])
    apply_style_to_excel(ws)
    adjust_column_width_and_row_height(ws)

    if output_path is None:
        output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not output_path:  # 如果用户取消选择
            print("未选择输出文件，程序退出。")
            sys.exit(1)

    wb.save(output_path)
    print(f"Excel文件已保存到: {output_path}")
    # 删除解压出的文件夹
    shutil.rmtree(extract_path)

if __name__ == "__main__":
    root = Tk()
    root.withdraw()  # 不显示Tk根窗口

    user_config = read_config()

    # 显示使用指南对话框
    if user_config.get('show_guide', True):
        show_guide = messagebox.askyesno(
            "雷腾小伙伴减负指南",

            "请将压缩包文件命名为以下格式：\n"

            "YYYY-MM-DD-文档名称-工作时长h.xxx\n"

            "例如:2024-01-31-质证意见第2版-1h.docx\n"

            "所有文件将被解压并整理到表格中，不符合命名规则的文件将被忽略。\n\n"
            
            "下次运行程序时是否还要显示这条消息？"
        )
        user_config['show_guide'] = show_guide
        write_config(user_config)

    # 选择要上传的压缩包
    messagebox.showinfo("选择压缩包", "请选择需要处理的压缩包文件。")
    zip_path = filedialog.askopenfilename(
        title="请选择压缩包",
        initialdir=os.path.expanduser("~"),
        filetypes=[("Zip files", "*.zip")]
    )
    if not zip_path:
        sys.exit("未选择压缩包，程序退出。")

    # 使用用户上次选择的路径作为默认路径
    default_path = user_config.get('default_path', os.path.join(os.path.expanduser("~"), "Desktop"))

    # 选择解压目标文件夹
    messagebox.showinfo("选择解压目标文件夹", "请选择要将临时内容解压到哪个文件夹。")
    extract_path = filedialog.askdirectory(
        title="选择解压目标文件夹",
        initialdir=default_path
    )
    if not extract_path:
        sys.exit("未选择解压目标文件夹，程序退出。")

    user_config['default_path'] = extract_path
    write_config(user_config)

    # 选择保存Excel文件的位置
    messagebox.showinfo("保存Excel文件", "请选择输出Excel文件的保存位置。")
    output_path = filedialog.asksaveasfilename(
        title="保存Excel文件",
        initialdir=default_path,
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not output_path:
        sys.exit("未选择输出文件，程序退出。")

    extract_and_process(zip_path, output_path)
