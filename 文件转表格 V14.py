import sys
import zipfile
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import os
import time
from tkinter import simpledialog
from tkinter import filedialog, messagebox, Tk, Toplevel, Label, Button, LEFT, RIGHT  
import chardet  # 用于编码检测
import shutil  # 用于删除文件和文件夹
import json  # 用于读写配置文件

config_file = 'user_config.json'

def read_config():
    """读取用户配置"""
    try:
        with open(config_file, 'r') as file:
            return json.load(file)
    except FileNotFoundError:
        return {}

def write_config(config):
    """写入用户配置"""
    with open(config_file, 'w') as file:
        json.dump(config, file)

# 设置环境变量以抑制Tkinter的弃用警告
os.environ['TK_SILENCE_DEPRECATION'] = '1'

def detect_encoding(byte_sequence):
    """检测字节序列的编码"""
    result = chardet.detect(byte_sequence)
    encoding = result['encoding']
    return encoding

def ensure_correct_encoding(file_name):
    """尝试以正确的编码解码文件名"""
    try:
        # 尝试使用utf-8解码
        decoded_name = file_name.encode('utf-8').decode('utf-8')
    except UnicodeDecodeError:
        try:
            # 如果失败，尝试使用系统默认编码
            decoded_name = file_name.encode('cp437').decode()
        except UnicodeDecodeError:
            # 如果仍然失败，保持原样
            decoded_name = file_name
    return decoded_name

def custom_dialog(root):
    """弹出自定义对话框让用户选择处理文件夹或压缩包"""
    dialog = Toplevel(root)
    dialog.title("选择处理类型")
    Label(dialog, text="你想要处理一个文件夹还是一个压缩包？").pack(pady=10)

    #对话框位置
    dialog_width = 200
    dialog_height = 100
    screen_width = dialog.winfo_screenwidth()
    screen_height = dialog.winfo_screenheight()
    x = (screen_width - dialog_width) / 2
    y = (screen_height - dialog_height) / 2
    dialog.geometry(f'{dialog_width}x{dialog_height}+{int(x)}+{int(y)}')


    def on_zip():
        global user_choice
        user_choice = '压缩包'
        dialog.destroy()

    def on_folder():
        global user_choice
        user_choice = '文件夹'
        dialog.destroy()

    Button(dialog, text="压缩包", command=on_zip).pack(side=LEFT, padx=10, pady=10)
    Button(dialog, text="文件夹", command=on_folder).pack(side=RIGHT, padx=10, pady=10)
    dialog.transient(root)  # Make the dialog on top of the main window
    dialog.grab_set()  # Make the dialog modal
    root.wait_window(dialog)  # Wait for the dialog to close before continue

def apply_style_to_excel(ws):
    """应用样式到Excel工作表"""
    # 设置首行的样式
    header_font = Font(name='微软雅黑', size=24, bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="6495ED", end_color="6495ED", fill_type="solid")

    for cell in ws["1:1"]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill

    # 应用样式到整个表格中的所有单元格
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    all_alignment = Alignment(wrap_text=True)
    normal_font = Font(size=16)
    
    # 从第2行开始，到倒数第二行
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row-1):
        for cell in row:
            cell.border = thin_border
            cell.alignment = all_alignment

            if cell.column == 2:  # 如果是第二列
            # 只为文档名称添加《》，如果它们还未被添加
               if not (str(cell.value).startswith('《') and str(cell.value).endswith('》')):
                    cell.value = f"《{cell.value}》" #加《》

            cell.font = normal_font
    
    # 设置最后一行的样式
    last_row_font = Font(size=24, bold=True)
    last_row_fill = PatternFill(start_color="D56A6B", end_color="D56A6B", fill_type="solid")
    last_row_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[ws.max_row]:
        cell.font = last_row_font
        cell.alignment = last_row_alignment
        cell.fill = last_row_fill
        cell.border = thin_border

    # 设置列宽
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 40

def process_file(root, file_path, file_pattern):
    """解析文件名以获取信息，若无创建日期或工作时长，则通过对话框提醒用户输入"""
    form_time = work_hours = doc_name = None

    # 尝试解析文件名获取文档日期、名称和工作时长
    match = re.match(file_pattern, os.path.basename(file_path))
    if match:
        groups = match.groups()
        form_time = groups[0]  # 文件日期
        doc_name = groups[1]  # 文件名称
        work_hours = float(groups[2])  # 文件工作时长
    else:
        # 获取文件的最后修改时间作为日期
        mod_time = time.localtime(os.path.getmtime(file_path))
        form_time = time.strftime("%Y-%m-%d", mod_time)
        
        # 获取文件名称，如果名称没有给出，询问用户
        doc_name = simpledialog.askstring("文档名称", f"请输入文件'{os.path.basename(file_path)}'的文档名称:", parent=root)
        if not doc_name:  # 如果用户取消或输入为空
            doc_name = "《未知文档》"
        
        # 获取文件工作时长，如果没有给出，询问用户
        work_hours = simpledialog.askfloat("工作时长", f"请输入文件'{os.path.basename(file_path)}'的工作时长（小时）:", parent=root)
        if work_hours is None:  # 如果用户取消
            work_hours = 0  # 可以设定一个默认值或者也可以选择退出

    return (form_time, doc_name, work_hours)

def excel_from_files(files_info):
    """从文件信息创建Excel工作簿"""
    wb = Workbook()
    ws = wb.active
    ws.append(["文件最后修改时间", "文档名称", "工作时长"])

    total_hours = 0
    file_count = 0  # 初始化文件数量计数
    for file_info in files_info:
        ws.append(file_info)
        total_hours += file_info[2]  # 累加工作时长
        file_count += 1  # 累加文件数量

    # 添加总计行，在第三列显示总工作时长，在第二列显示文件总数
    ws.append(["总计", file_count, f"{total_hours}h"])
    apply_style_to_excel(ws)
    return wb

def process_directory(directory_path, output_path):
    """处理指定文件夹内的文件，然后形成Excel表格"""
    # 初始化保存处理后文件信息的列表
    files_info = []
    
    # 遍历目录内的所有文件
    for root_dir, dirs, files in os.walk(directory_path):
        for file_name in files:
            # 忽略隐藏文件和系统文件
            if file_name.startswith('.') or file_name.startswith('__MACOSX') or file_name == 'DS_Store':
                continue

            # 组合成完整的文件路径
            file_path = os.path.join(root_dir, file_name)
            # 调用process_file函数处理文件，并获取文件信息
            file_info = process_file(file_path)
            # 如果得到了有效文件信息，则添加到信息列表中
            if file_info is not None:
                files_info.append(file_info)

    # 如果文件信息列表为空，则提示用户并返回
    if not files_info:
        messagebox.showinfo("提示", "未找到可处理的文件。")
        return

    # 调用函数，将收集到的文件信息转换为Excel工作簿
    wb = excel_from_files(files_info)
    # 保存Excel工作簿到用户指定的输出路径
    wb.save(output_path)
    messagebox.showinfo("完成", f"Excel文件已保存到: {output_path}")

def extract_and_process(zip_path, output_path=None):
    user_config = read_config()
    messagebox.showinfo("选择压缩包", "请选择包含需要处理的文件的压缩包文件。")
    zip_path = filedialog.askopenfilename(title="选择压缩包", initialdir=user_config.get('last_zip', os.path.expanduser("~")), filetypes=[("Zip files", "*.zip")])
    if zip_path:
        # 保存用户选择的压缩包路径
        user_config['last_zip'] = zip_path
        write_config(user_config)
    """处理压缩包内的文件，然后形成表格"""
    file_pattern = r"(\d{4}-\d{1,2}-\d{1,2})-([\u4e00-\u9fa5\w]+)-(\d+(\.\d{1,2})?)h\.(docx|doc|xlsx|xls)"
    
    # 未提供输出路径，弹窗询问用户选择
    if output_path is None:
        messagebox.showinfo("保存Excel", "请选择一个位置来保存生成的Excel文件。")
        output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not output_path:  # 如果用户取消选择
            sys.exit(1)
    
    messagebox.showinfo("解压目标文件夹", "请选择一个临时缓存文件夹以解压缩包内容。")
    extract_path = filedialog.askdirectory(title="请选择解压目标文件夹")
    if not extract_path:
        sys.exit("未选择解压目标文件夹，程序退出。")

    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        zip_ref.extractall(extract_path)

    files_info = []
    for root, dirs, files in os.walk(extract_path):
        for file in files:
            if file.startswith('.') or file.startswith('__MACOSX') or file == 'DS_Store':
                continue
            file_path = os.path.join(root, file)
            file_info = process_file(root, file_path, file_pattern)
            if file_info is not None:
                files_info.append(file_info)

    shutil.rmtree(extract_path) #删除临时文件

    wb = excel_from_files(files_info)
    wb.save(output_path)
    print(f"Excel文件已保存到: {output_path}")

def main():
    root = Tk()
    root.withdraw()  # 隐藏主窗口

    user_config = read_config()
    
    # 显示使用指南对话框
    if user_config.get('show_guide', True):
        show_guide = messagebox.askyesno(
            "使用指南",
            "请将文件命名为以下格式：\n"
            "\n"
            "YYYY-MM-DD-文档名称-工作时长h.xxx\n"
            "例如: 2024-01-31-质证意见第2版-1h.docx\n"
            "\n"
            "文件会被整理到表格中。\n"
            "没有写名字的会提示填写\n"
            "没有写工作时长的文件会提示填写。\n"
            "没有写时间的文件会使用最后修改时间。\n"
            "\n"
            "下次运行程序时是否还要显示这条消息？",
            parent=root
        )
        user_config['show_guide'] = not show_guide
        write_config(user_config)
    
    # 调用自定义对话框让用户选择处理文件夹或压缩包
    custom_dialog(root)
    
    if user_choice == '文件夹':
        directory_path = filedialog.askdirectory(
            title="选择文件夹",
            initialdir=user_config.get('last_directory', os.path.expanduser("~")),
            parent=root
        )
        if directory_path:
            user_config['last_directory'] = directory_path
            write_config(user_config)
            output_path = filedialog.asksaveasfilename(
                title="保存Excel文件",
                initialdir=directory_path,
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                parent=root
            )
            if output_path:
                process_directory(directory_path, output_path, root)
            else:
                messagebox.showinfo("提示", "没有选择输出文件，操作已取消。", parent=root)
        else:
            messagebox.showinfo("提示", "没有选择文件夹，操作已取消。", parent=root)

    elif user_choice == '压缩包':
        zip_path = filedialog.askopenfilename(
            title="选择压缩包",
            initialdir=user_config.get('last_zip', os.path.expanduser("~")),
            filetypes=[("Zip files", "*.zip")],
            parent=root
        )
        if zip_path:
            user_config['last_zip'] = zip_path
            write_config(user_config)
            output_path = filedialog.asksaveasfilename(
                title="保存Excel文件",
                initialdir=os.path.dirname(zip_path),
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                parent=root
            )
            if output_path:
                extract_and_process(zip_path, output_path)
            else:
                messagebox.showinfo("提示", "没有选择输出文件，操作已取消。", parent=root)
        else:
            messagebox.showinfo("提示", "没有选择压缩包，操作已取消。", parent=root)

if __name__ == "__main__":
    main()

    

